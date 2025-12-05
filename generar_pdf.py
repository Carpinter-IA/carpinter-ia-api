# -*- coding: utf-8 -*-
"""
generar_pdf.py
Módulo auxiliar simple para Carpinter-IA que genera un PDF en memoria a partir
de la lista de piezas + metadatos.

Funciones públicas (cualquiera de estas puede ser llamada por app.py):
- generar_pdf_bytes(piezas, meta=None, material=None, espesor=None, cliente=None) -> bytes
- generar_pdf(piezas, meta=None, material=None, espesor=None, cliente=None) -> bytes
- create_pdf_bytes(...), generate_pdf_bytes(...), exportar_pdf_maestro_bytes(...)

También exporta generar_xlsx_bytes(piezas, ...) por si la API lo solicita.

Dependencias: reportlab, openpyxl (ya las tienes en requirements).
"""

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io
import datetime
import base64

# XLSX
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def _normalize_piezas(piezas):
    """
    Asegura que piezas es lista de dicts con keys (cantidad,largo,ancho,ocr_texto)
    """
    out = []
    if not piezas:
        return out
    for p in piezas:
        try:
            cant = int(p.get("cantidad", p.get("qty", 1)))
        except Exception:
            cant = 1
        try:
            largo = int(p.get("largo", p.get("length", 0)))
        except Exception:
            largo = 0
        try:
            ancho = int(p.get("ancho", p.get("width", 0)))
        except Exception:
            ancho = 0
        ocr_texto = p.get("ocr_texto", f"{cant} {largo}x{ancho}")
        out.append({"cantidad": cant, "largo": largo, "ancho": ancho, "ocr_texto": ocr_texto})
    return out

def generar_pdf_bytes(piezas, meta=None, material=None, espesor=None, cliente=None):
    """
    Devuelve bytes con el PDF creado.
    Simple: cabecera con metadatos + tabla con piezas.
    """
    piezas = _normalize_piezas(piezas or [])
    buf = io.BytesIO()
    # documento apaisado (mejor para tablas largas)
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)

    styles = getSampleStyleSheet()
    elems = []

    # Cabecera
    titulo = Paragraph("<b>Carpinter-IA — Despiece</b>", styles["Title"])
    elems.append(titulo)
    elems.append(Spacer(1, 4*mm))

    # metadatos pequeños
    info_lines = []
    fecha = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    info_lines.append(f"Fecha: {fecha}")
    if cliente:
        info_lines.append(f"Cliente/Proyecto: {cliente}")
    if material:
        info_lines.append(f"Material: {material}")
    if espesor:
        info_lines.append(f"Espesor: {espesor}")
    if meta and isinstance(meta, dict) and meta.get("image_path"):
        info_lines.append(f"Imagen: {meta.get('image_path')}")
    for ln in info_lines:
        elems.append(Paragraph(ln, styles["Normal"]))
    elems.append(Spacer(1, 6*mm))

    # Tabla de piezas
    data = [["#", "Cantidad", "Largo (mm)", "Ancho (mm)", "OCR texto"]]
    for i, p in enumerate(piezas, 1):
        data.append([str(i), str(p["cantidad"]), str(p["largo"]), str(p["ancho"]), p.get("ocr_texto", "")])

    if len(data) == 1:
        elems.append(Paragraph("No se detectaron piezas.", styles["Normal"]))
    else:
        col_widths = [20*mm, 30*mm, 35*mm, 35*mm, 130*mm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        tblstyle = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#e8e8e8")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ])
        table.setStyle(tblstyle)
        elems.append(table)

    elems.append(Spacer(1,6*mm))
    elems.append(Paragraph("Generado por Carpinter-IA", styles["Italic"]))

    doc.build(elems)
    pdf = buf.getvalue()
    buf.close()
    return pdf

# alias y compatibilidades
def generate_pdf_bytes(*args, **kwargs):
    return generar_pdf_bytes(*args, **kwargs)

def generar_pdf(*args, **kwargs):
    return generar_pdf_bytes(*args, **kwargs)

def create_pdf_bytes(*args, **kwargs):
    return generar_pdf_bytes(*args, **kwargs)

def exportar_pdf_maestro_bytes(*args, **kwargs):
    return generar_pdf_bytes(*args, **kwargs)

# --- XLSX generator (opcional) ---
def generar_xlsx_bytes(piezas, meta=None, material=None, espesor=None, cliente=None):
    piezas = _normalize_piezas(piezas or [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Despiece"
    headers = ["#", "Cantidad", "Largo (mm)", "Ancho (mm)", "OCR texto"]
    ws.append(headers)
    for i, p in enumerate(piezas, 1):
        ws.append([i, p["cantidad"], p["largo"], p["ancho"], p.get("ocr_texto","")])
    # ajustar anchos basicos
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18 if i < 5 else 40
    bio = io.BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    bio.close()
    return data

def generate_xlsx_bytes(*a, **k):
    return generar_xlsx_bytes(*a, **k)

# --- helpers para base64 (si app.py lo necesita) ---
def generar_pdf_base64(piezas, meta=None, material=None, espesor=None, cliente=None):
    pdf = generar_pdf_bytes(piezas, meta=meta, material=material, espesor=espesor, cliente=cliente)
    return base64.b64encode(pdf).decode("ascii")

def generar_xlsx_base64(piezas, meta=None, material=None, espesor=None, cliente=None):
    x = generar_xlsx_bytes(piezas, meta=meta, material=material, espesor=espesor, cliente=cliente)
    return base64.b64encode(x).decode("ascii")
